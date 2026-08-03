#!/usr/bin/env python3
"""
update_tracker.py — import new jobs into applications.xlsx.
Usage:
    python3 update_tracker.py --import-csv /tmp/jobs_combined.csv
    python3 update_tracker.py --add --title "DevOps Engineer" --company "Amazon" --link "https://..."
    python3 update_tracker.py --set-status "Applied" --row 3
"""

import sys
import csv
import argparse
from datetime import date
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent
sys.path.insert(0, str(SCRIPT_DIR))
import config

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("[ERROR] openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

TODAY = date.today().isoformat()

COLUMNS = [
    "Date Added", "Source", "Job Title", "Company", "Location",
    "Salary", "Link", "Status", "Applied Date", "Notes"
]

# Status colours (hex, no #)
STATUS_COLORS = {
    "Wishlist":   "FFF9C4",   # light yellow
    "Applied":   "BBDEFB",   # light blue
    "Interview":  "C8E6C9",   # light green
    "Offer":     "A5D6A7",   # medium green
    "Rejected":  "FFCDD2",   # light red
}

HDR_FILL   = PatternFill("solid", fgColor="1565C0")   # dark blue
HDR_FONT   = Font(color="FFFFFF", bold=True)
thin       = Side(style="thin", color="BDBDBD")
BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)


def ensure_workbook(path: Path):
    """Create workbook with headers if it doesn't exist."""
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(COLUMNS)
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font   = HDR_FONT
            cell.fill   = HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

        # Column widths
        widths = [12, 18, 30, 22, 20, 15, 45, 12, 14, 35]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        ws.freeze_panes = "A2"
        wb.save(path)
        print(f"[NEW] Created workbook at {path}")
    return wb


def import_csv(path: Path, wb):
    """Read CSV and append new rows, skipping duplicates by link."""
    ws = wb.active

    # Collect existing links
    existing_links = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        link = row[6] if len(row) > 6 else ""
        if link:
            existing_links.add(link)

    reader = csv.DictReader(path.read_text(encoding="utf-8").splitlines())
    added = 0
    for job in reader:
        link = job.get("link", "").strip()
        if link and link in existing_links:
            print(f"  [SKIP] Duplicate: {job.get('title','')}")
            continue

        row_num = ws.max_row + 1
        values = [
            TODAY,
            job.get("source", ""),
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            job.get("salary", ""),
            link,
            "Wishlist",   # default status
            "",           # applied date
            "",           # notes
        ]
        ws.append(values)

        # Colour the status cell
        status_cell = ws.cell(row=row_num, column=8)
        _colour_status_cell(status_cell, "Wishlist")

        # Wrap text for long cells
        for col in range(1, len(COLUMNS) + 1):
            ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True)
            ws.cell(row=row_num, column=col).border = BORDER

        added += 1
        print(f"  [ADD] {job.get('title','')} @ {job.get('company','')}")

    wb.save(config.EXCEL_PATH)
    print(f"\n✓ Added {added} new jobs → {config.EXCEL_PATH}")
    return added


def _colour_status_cell(cell, status: str):
    color = STATUS_COLORS.get(status, "FFFFFF")
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center")


def set_status(wb, path: Path, row: int, new_status: str):
    """Update status column for a specific row."""
    ws = wb.active
    if row < 2 or row > ws.max_row:
        print(f"[ERROR] Row {row} out of range")
        return
    cell = ws.cell(row=row, column=8)
    cell.value = new_status
    _colour_status_cell(cell, new_status)
    wb.save(path)
    print(f"[ROW {row}] Status → {new_status}")


def print_usage():
    print("""Usage:
  python3 update_tracker.py --import-csv /path/to/jobs.csv
  python3 update_tracker.py --add --title "DevOps" --company "Acme" --link "https://..."
  python3 update_tracker.py --set-status Applied --row 3
  python3 update_tracker.py --list
""")


def cmd_add(args):
    wb = ensure_workbook(config.EXCEL_PATH)
    ws = wb.active
    row_num = ws.max_row + 1
    values = [TODAY, args.source or "", args.title, args.company or "",
             args.location or "", args.salary or "", args.link or "",
             "Wishlist", "", args.notes or ""]
    ws.append(values)
    for col in range(1, len(COLUMNS) + 1):
        ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_num, column=col).border = BORDER
    _colour_status_cell(ws.cell(row=row_num, column=8), "Wishlist")
    wb.save(config.EXCEL_PATH)
    print(f"[ADD] Row {row_num}: {args.title}")


def cmd_list(wb):
    ws = wb.active
    print(f"\n{'Row':<4} {'Date':<12} {'Status':<10} {'Title':<35} {'Company'}")
    print("-" * 100)
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if row[0]:  # skip empty rows
            print(f"{i:<4} {str(row[0]):<12} {str(row[7]):<10} {str(row[2]):<35} {str(row[3])}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--import-csv", "-i")
    parser.add_argument("--add", action="store_true")
    parser.add_argument("--title", "-t")
    parser.add_argument("--company", "-c")
    parser.add_argument("--location", "-l", default="")
    parser.add_argument("--salary", "-s", default="")
    parser.add_argument("--link", "-u", default="")
    parser.add_argument("--notes", "-n", default="")
    parser.add_argument("--source", default="Manual")
    parser.add_argument("--set-status")
    parser.add_argument("--row", type=int)
    parser.add_argument("--list", action="store_true")
    args = parser.parse_args()

    if not any([args.import_csv, args.add, args.set_status, args.list]):
        print_usage()
        return

    config.EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = ensure_workbook(config.EXCEL_PATH)

    if args.import_csv:
        import_csv(Path(args.import_csv), wb)

    if args.add:
        cmd_add(args)

    if args.set_status and args.row:
        set_status(wb, config.EXCEL_PATH, args.row, args.set_status)

    if args.list:
        cmd_list(wb)


if __name__ == "__main__":
    main()
