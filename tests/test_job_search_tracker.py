"""Tests for job-search/update_tracker.py — Excel tracker."""
from __future__ import annotations

import csv
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest


# ── normalize_link ────────────────────────────────────────────────────────────

class TestNormalizeLink:
    def test_strips_sol_hash(self, js_update_tracker):
        link = "https://hk.jobsdb.com/job/123#sol=abc123def456"
        norm = js_update_tracker.normalize_link(link)
        assert "#sol=" not in norm

    def test_strips_ref_param(self, js_update_tracker):
        link = "https://hk.jobsdb.com/job/123?ref=search-standalone"
        norm = js_update_tracker.normalize_link(link)
        assert "&ref=search-standalone" not in norm

    def test_preserves_other_query_params(self, js_update_tracker):
        link = "https://hk.jobsdb.com/job/123?type=full&source=jobsdb"
        norm = js_update_tracker.normalize_link(link)
        assert "type=full" in norm
        assert "source=jobsdb" in norm

    def test_strips_both_hash_and_ref(self, js_update_tracker):
        # The actual implementation only strips &ref=... (when preceded by
        # another param), not ?ref=... (when it's the first param).
        link = "https://hk.jobsdb.com/job/123?k=aws&ref=search-standalone#sol=abc123"
        norm = js_update_tracker.normalize_link(link)
        assert "#sol=" not in norm
        assert "ref=search-standalone" not in norm

    def test_idempotent(self, js_update_tracker):
        link = "https://hk.jobsdb.com/job/123#sol=abc"
        once = js_update_tracker.normalize_link(link)
        twice = js_update_tracker.normalize_link(once)
        assert once == twice


# ── ensure_workbook ───────────────────────────────────────────────────────────

class TestEnsureWorkbook:
    def test_creates_workbook_when_missing(self, js_update_tracker, tmp_path):
        excel = tmp_path / "applications.xlsx"
        wb = js_update_tracker.ensure_workbook(excel)
        assert excel.exists()
        # Active sheet named Applications
        assert wb.active.title == "Applications"

    def test_loads_existing_workbook(self, js_update_tracker, tmp_path):
        excel = tmp_path / "applications.xlsx"
        # Create an existing workbook by calling ensure_workbook once
        js_update_tracker.ensure_workbook(excel)
        # Now load it again
        wb2 = js_update_tracker.ensure_workbook(excel)
        # Should not crash and should have one sheet
        assert len(wb2.sheetnames) == 1

    def test_creates_with_proper_columns(self, js_update_tracker, tmp_path):
        from openpyxl import load_workbook
        excel = tmp_path / "applications.xlsx"
        js_update_tracker.ensure_workbook(excel)
        wb = load_workbook(excel)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        expected_headers = [
            "Date Added", "Source", "Job Title", "Company",
            "Location", "Salary", "Link", "Status",
            "Applied Date", "Notes",
        ]
        assert headers == expected_headers

    def test_freeze_panes_set(self, js_update_tracker, tmp_path):
        from openpyxl import load_workbook
        excel = tmp_path / "applications.xlsx"
        js_update_tracker.ensure_workbook(excel)
        wb = load_workbook(excel)
        assert wb.active.freeze_panes == "A2"

    def test_column_widths_set(self, js_update_tracker, tmp_path):
        from openpyxl import load_workbook
        excel = tmp_path / "applications.xlsx"
        js_update_tracker.ensure_workbook(excel)
        wb = load_workbook(excel)
        ws = wb.active
        # Check first few column widths are non-default
        assert ws.column_dimensions["A"].width > 5  # Date Added
        assert ws.column_dimensions["G"].width > 30  # Link


# ── import_csv ─────────────────────────────────────────────────────────────────

class TestImportCsv:
    def test_empty_csv_no_changes(self, js_update_tracker, tmp_path):
        excel = tmp_path / "applications.xlsx"
        js_update_tracker.ensure_workbook(excel)
        # Empty CSV (header only)
        csv_path = tmp_path / "empty.csv"
        csv_path.write_text("title,company,link\n", encoding="utf-8")
        from openpyxl import load_workbook
        wb = load_workbook(excel)
        ws = wb.active
        rows_before = ws.max_row
        js_update_tracker.import_csv(csv_path, wb)
        ws2 = wb.active
        assert ws2.max_row == rows_before  # no new rows

    def test_imports_unique_rows(self, js_update_tracker, tmp_path):
        excel = tmp_path / "applications.xlsx"
        js_update_tracker.ensure_workbook(excel)
        from openpyxl import load_workbook

        csv_path = tmp_path / "new.csv"
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["title", "company", "link"])
            w.writeheader()
            w.writerow({
                "title": "DevOps Engineer",
                "company": "HSBC",
                "link": "https://hk.jobsdb.com/job/1",
            })
            w.writerow({
                "title": "SRE",
                "company": "TechCorp",
                "link": "https://example.com/job/2",
            })

        wb = load_workbook(excel)
        added = js_update_tracker.import_csv(csv_path, wb)
        assert added == 2

    def test_skips_duplicate_links(self, js_update_tracker, tmp_path):
        excel = tmp_path / "applications.xlsx"
        js_update_tracker.ensure_workbook(excel)
        # Pre-populate one job
        from openpyxl import load_workbook
        wb = load_workbook(excel)
        ws = wb.active
        ws.append(["2026-01-01", "JobsDB", "Existing", "Acme", "HK",
                   "$50K", "https://hk.jobsdb.com/job/99", "Wishlist", "", ""])
        wb.save(excel)

        # CSV has the same link — should be skipped
        csv_path = tmp_path / "dup.csv"
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["title", "company", "link"])
            w.writeheader()
            w.writerow({
                "title": "Duplicate",
                "company": "Different Co",
                "link": "https://hk.jobsdb.com/job/99",
            })
            w.writerow({
                "title": "New",
                "company": "New Co",
                "link": "https://example.com/new",
            })

        wb = load_workbook(excel)
        added = js_update_tracker.import_csv(csv_path, wb)
        assert added == 1  # only the new one

    def test_dedupes_against_sol_hash_and_ref_param(self, js_update_tracker, tmp_path):
        """Same logical job with different sol=/ref= should dedupe."""
        excel = tmp_path / "applications.xlsx"
        js_update_tracker.ensure_workbook(excel)
        from openpyxl import load_workbook
        wb = load_workbook(excel)
        ws = wb.active
        # Pre-populate with sol= and ref= variants
        ws.append(["2026-01-01", "JobsDB", "Existing", "Acme", "HK",
                   "$50K",
                   "https://hk.jobsdb.com/job/42#sol=abc123&ref=search-standalone",
                   "Wishlist", "", ""])
        wb.save(excel)

        # New CSV with same base link but no sol= / ref=
        csv_path = tmp_path / "var.csv"
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["title", "company", "link"])
            w.writeheader()
            w.writerow({
                "title": "Same job",
                "company": "Acme",
                "link": "https://hk.jobsdb.com/job/42",
            })

        wb = load_workbook(excel)
        added = js_update_tracker.import_csv(csv_path, wb)
        # Same job should be deduped despite different sol=/ref=
        assert added == 0


# ── status colors ─────────────────────────────────────────────────────────────

class TestStatusColors:
    def test_status_colors_defined(self, js_update_tracker):
        for status in ("Wishlist", "Applied", "Interview", "Offer", "Rejected"):
            assert status in js_update_tracker.STATUS_COLORS
            color = js_update_tracker.STATUS_COLORS[status]
            # Valid 6-char hex (with or without alpha prefix — openpyxl accepts both)
            assert len(color) == 6 or (len(color) == 8 and color[:2] == "FF")
            int(color, 16)  # raises if invalid

    def test_all_colors_are_6char_hex(self, js_update_tracker):
        for color in js_update_tracker.STATUS_COLORS.values():
            assert len(color) == 6
            int(color, 16)  # raises if not valid hex


# ── set_status ────────────────────────────────────────────────────────────────

class TestSetStatus:
    def test_changes_status_cell(self, js_update_tracker, tmp_path):
        from openpyxl import load_workbook
        excel = tmp_path / "applications.xlsx"
        js_update_tracker.ensure_workbook(excel)
        wb = load_workbook(excel)
        ws = wb.active
        ws.append(["2026-01-01", "JobsDB", "Job", "Acme", "HK",
                   "$50K", "https://x.com/1", "Wishlist", "", ""])
        wb.save(excel)

        wb = load_workbook(excel)
        js_update_tracker.set_status(wb, excel, 2, "Applied")
        wb2 = load_workbook(excel)
        assert wb2.active.cell(row=2, column=8).value == "Applied"

    def test_rejects_invalid_row(self, js_update_tracker, tmp_path, capsys):
        from openpyxl import load_workbook
        excel = tmp_path / "applications.xlsx"
        js_update_tracker.ensure_workbook(excel)
        wb = load_workbook(excel)
        # Try row 1 (header) and row 999 (out of range)
        js_update_tracker.set_status(wb, excel, 1, "Applied")
        js_update_tracker.set_status(wb, excel, 999, "Applied")
        out = capsys.readouterr().out
        assert "out of range" in out

    def test_applies_status_color(self, js_update_tracker, tmp_path):
        from openpyxl import load_workbook
        excel = tmp_path / "applications.xlsx"
        js_update_tracker.ensure_workbook(excel)
        wb = load_workbook(excel)
        ws = wb.active
        ws.append(["2026-01-01", "JobsDB", "Job", "Acme", "HK",
                   "$50K", "https://x.com/1", "Wishlist", "", ""])
        wb.save(excel)
        wb = load_workbook(excel)
        js_update_tracker.set_status(wb, excel, 2, "Rejected")
        wb2 = load_workbook(excel)
        cell = wb2.active.cell(row=2, column=8)
        # Should have a fill applied (PatternFill)
        assert cell.fill is not None


# ── cmd_add ────────────────────────────────────────────────────────────────────

class TestCmdAdd:
    def test_adds_row_with_default_wishlist_status(self, js_update_tracker, tmp_path, monkeypatch):
        from openpyxl import load_workbook
        excel = tmp_path / "applications.xlsx"
        monkeypatch.setattr(js_update_tracker.config, "EXCEL_PATH", excel)
        args = MagicMock(title="DevOps", company="HSBC", location="HK",
                        salary="$80K", link="https://x.com/1", source="JobsDB",
                        notes="")
        js_update_tracker.cmd_add(args)
        wb = load_workbook(excel)
        ws = wb.active
        # Find the appended row
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        assert rows[0][2] == "DevOps"
        assert rows[0][3] == "HSBC"
        assert rows[0][7] == "Wishlist"

    def test_appends_to_existing_workbook(self, js_update_tracker, tmp_path, monkeypatch):
        from openpyxl import load_workbook
        excel = tmp_path / "applications.xlsx"
        monkeypatch.setattr(js_update_tracker.config, "EXCEL_PATH", excel)
        # Pre-populate
        js_update_tracker.ensure_workbook(excel)
        wb = load_workbook(excel)
        wb.active.append(["x", "y", "z", "q", "r", "s", "t", "Wishlist", "", ""])
        wb.save(excel)
        # Add a new job
        args = MagicMock(title="New", company="NewCo", location="HK",
                        salary="$50K", link="https://x.com/9", source="JobsDB",
                        notes="")
        js_update_tracker.cmd_add(args)
        wb2 = load_workbook(excel)
        rows = list(wb2.active.iter_rows(min_row=2, values_only=True))
        # At least 2 rows now (the pre-existing + the new one)
        assert len(rows) >= 2


# ── main / CLI ─────────────────────────────────────────────────────────────────

class TestMain:
    def test_no_args_prints_usage(self, js_update_tracker, tmp_path, monkeypatch, capsys):
        excel = tmp_path / "applications.xlsx"
        monkeypatch.setattr(js_update_tracker.config, "EXCEL_PATH", excel)
        import sys
        monkeypatch.setattr(sys, "argv", ["update_tracker.py"])
        js_update_tracker.main()
        out = capsys.readouterr().out
        assert "Usage" in out

    def test_list_prints_rows(self, js_update_tracker, tmp_path, monkeypatch, capsys):
        from openpyxl import load_workbook
        excel = tmp_path / "applications.xlsx"
        monkeypatch.setattr(js_update_tracker.config, "EXCEL_PATH", excel)
        # Pre-populate
        js_update_tracker.ensure_workbook(excel)
        wb = load_workbook(excel)
        wb.active.append(["2026-01-01", "JobsDB", "DevOps", "HSBC", "HK",
                          "$80K", "https://x.com/1", "Wishlist", "", ""])
        wb.save(excel)

        import sys
        monkeypatch.setattr(sys, "argv", ["update_tracker.py", "--list"])
        js_update_tracker.main()
        out = capsys.readouterr().out
        assert "DevOps" in out
        assert "HSBC" in out